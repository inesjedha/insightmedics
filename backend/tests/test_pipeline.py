"""Tests du pipeline d'audit : une mini-base aux défauts connus doit les déclencher tous."""

import io
from pathlib import Path

import pandas as pd
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.pipeline.ingest import IngestError, ingest
from app.pipeline.runner import run_audit


@pytest.fixture()
def dirty_csv(tmp_path: Path) -> Path:
    """Base volontairement sale : doublon strict, doublon d'ID, code 999,
    colonne vide, colonne constante, valeur extrême, PII téléphone."""
    df = pd.DataFrame({
        "matricule": ["P1", "P2", "P3", "P3", "P4", "P4"],
        "age": [34, 45, 999, 999, 51, 51],
        "poids": [70, 82, 65, 65, 400, 400],       # 400 kg = outlier
        "centre": ["Tunis"] * 6,                    # constante
        "commentaire": [None] * 6,                  # vide
        "tel_patient": ["+21620304050"] * 6,        # PII
        "crp": [12.0, None, 8.0, 8.0, None, None],  # manquants
    })
    # P4 dupliqué strictement, P3 dupliqué aussi
    p = tmp_path / "dirty.csv"
    df.to_csv(p, index=False)
    return p


def test_run_audit_detecte_les_defauts(dirty_csv: Path):
    r = run_audit(dirty_csv, "dirty.csv")
    p = r["profiling"]
    s = p["structure"]

    assert r["row_count"] == 6 and r["column_count"] == 7
    assert s["id_column"] == "matricule"
    assert s["duplicates"]["exact_rows"] >= 2          # lignes dupliquées
    assert s["duplicates"]["duplicate_ids"] >= 1       # ID en double
    assert "commentaire" in s["empty_columns"]
    assert "centre" in s["constant_columns"]

    age = next(c for c in p["columns"] if c["name"] == "age")
    assert 999 in age["suspected_missing_codes"]

    assert any(c["column"] == "tel_patient" for c in p["pii_candidates"])
    assert 0 <= r["score"] <= 100
    assert any(i["level"] == "critical" for i in r["issues"])


def test_ingest_refuse_les_formats_inconnus(tmp_path: Path):
    bad = tmp_path / "donnees.txt"
    bad.write_text("bonjour")
    with pytest.raises(IngestError):
        ingest(bad)


def test_api_upload_et_lecture(dirty_csv: Path):
    client = TestClient(app)
    with open(dirty_csv, "rb") as f:
        resp = client.post("/audit/upload", files={"file": ("dirty.csv", f, "text/csv")})
    assert resp.status_code == 200
    body = resp.json()
    assert body["rowCount"] == 6
    assert body["needsHumanReview"] is True            # anomalies critiques présentes

    audit_id = body["id"]
    assert client.get(f"/audit/{audit_id}").status_code == 200
    events = client.get(f"/audit/{audit_id}/events").json()
    assert any("Parsing OK" in e["message"] for e in events)


def test_api_refuse_mauvais_format():
    client = TestClient(app)
    resp = client.post("/audit/upload",
                       files={"file": ("x.txt", io.BytesIO(b"abc"), "text/plain")})
    assert resp.status_code == 415


# ---------------------------------------------------------------- livrables M6/M8

def _upload(client: TestClient, path: Path) -> str:
    with open(path, "rb") as f:
        resp = client.post("/audit/upload", files={"file": ("dirty.csv", f, "text/csv")})
    assert resp.status_code == 200
    return resp.json()["id"]


def test_api_classeur_xlsx_bout_en_bout(dirty_csv: Path):
    """Le classeur téléchargé contient les 10 onglets, dont Base analyse et Source
    anonymisée produits par le nettoyage M8 branché sur le fichier source persisté."""
    from openpyxl import load_workbook

    client = TestClient(app)
    audit_id = _upload(client, dirty_csv)
    resp = client.get(f"/audit/{audit_id}/workbook.xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Base analyse" in wb.sheetnames
    assert "Source anonymisée" in wb.sheetnames
    # Onglet Base analyse : id_anonyme en tête, PII et colonne vide retirées
    base = wb["Base analyse"]
    entetes = [c.value for c in next(base.iter_rows(max_row=1))]
    assert "id_anonyme" in entetes and "ligne_source" in entetes
    assert "tel_patient" not in entetes      # PII exclue
    assert "commentaire" not in entetes      # colonne vide exclue


def test_api_base_analyse_csv_anonymisee(dirty_csv: Path):
    client = TestClient(app)
    audit_id = _upload(client, dirty_csv)
    resp = client.get(f"/audit/{audit_id}/base_analyse.csv")
    assert resp.status_code == 200
    texte = resp.content.decode("utf-8")
    entete = texte.splitlines()[0]
    assert "id_anonyme" in entete and "ligne_source" in entete
    assert "tel_patient" not in entete       # PII jamais dans la base d'analyse
    assert "P001" in texte                   # identifiant anonyme généré


def test_api_base_analyse_sav_relisible(dirty_csv: Path, tmp_path: Path):
    import pyreadstat

    client = TestClient(app)
    audit_id = _upload(client, dirty_csv)
    resp = client.get(f"/audit/{audit_id}/base_analyse.sav")
    assert resp.status_code == 200
    out = tmp_path / "out.sav"
    out.write_bytes(resp.content)
    df, _ = pyreadstat.read_sav(str(out))    # doit se relire sans erreur
    assert "id_anonyme" in df.columns and len(df) == 6


def test_api_rapport_docx(dirty_csv: Path):
    from docx import Document

    client = TestClient(app)
    audit_id = _upload(client, dirty_csv)
    resp = client.get(f"/audit/{audit_id}/report.docx")
    assert resp.status_code == 200
    assert "wordprocessingml" in resp.headers["content-type"]
    doc = Document(io.BytesIO(resp.content))
    headings = [p.text for p in doc.paragraphs if p.style.name == "Heading 1"]
    assert len(headings) == 11               # 11 sections comme Hamza
