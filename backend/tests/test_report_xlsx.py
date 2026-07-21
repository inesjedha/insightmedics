"""Tests du classeur Excel d'audit (M6) — construction déterministe, sans API."""

import io

from openpyxl import load_workbook

from app.pipeline.report_xlsx import build_workbook
from app.pipeline.score_engine import compute_score
from tests.test_score_engine import CLEAN, FULL_SI

SYNTH_AI = {
    "study": {
        "primary_objective": "Évaluer l'évolution musculaire",
        "secondary_objectives": ["Mortalité à J30"],
        "primary_endpoint": {"description": "Épaisseur du quadriceps",
                             "candidate_columns": ["age"], "confidence": "medium"},
    },
    "dictionary": [
        {"name": "age", "meaning": "Âge du patient", "analytic_role": "confounder",
         "expected_unit": "années", "ambiguities": None},
    ],
    "assessment": {
        "findings": [
            {"id": "F001", "anomaly_class": "A", "severity": "critique", "certainty": "certain",
             "column": "age", "observed": "999", "rule_violated": "borne max 120",
             "title_fr": "Code 999", "explanation_fr": "…", "proposed_correction": "recode",
             "requires_source_verification": False},
            {"id": "F002", "anomaly_class": "C", "severity": "mineure", "certainty": "possible",
             "column": "sexe", "observed": None, "rule_violated": None,
             "title_fr": "Modalité rare", "explanation_fr": "…", "proposed_correction": None,
             "requires_source_verification": True},
        ],
        "cleaning_plan": [
            {"op_id": "op1", "operation": "recode_missing", "column": "age",
             "params": {}, "rationale_fr": "999 = manquant", "auto_safe": True},
        ],
        "client_decisions": [
            {"question_fr": "Fusionner les doublons ?", "column": "matricule",
             "options": ["oui", "non"], "consequence_fr": "…", "recommendation_fr": "vérifier"},
        ],
        "exploitability_verdict": {"level": 2, "label": "exploitable avec réserves",
                                   "justification_fr": "…"},
        "report_sections_fr": {"limites": "…", "plan_action": "…",
                               "plan_analyse_conditionnel": "Étape 1 : décrire la population…"},
        "pii_assessment": [],
    },
}

EXPECTED_SHEETS = ["Résumé audit", "Registre anomalies", "Dictionnaire",
                   "Concordance protocole", "Journal corrections", "Variables exclues",
                   "Décisions à valider", "Plan analyse"]


def _profiling_min():
    # profiling minimal cohérent avec le schéma attendu par le classeur
    return {
        "file": {"file_name": "test.sav"},
        "structure": CLEAN["structure"],
        "missing_summary": CLEAN["missing_summary"],
        "columns": [
            {"name": "matricule", "spss_label": "ID", "stat_type_guess": "nominal",
             "n": 100, "n_missing": 0, "pct_missing": 0.0, "n_distinct": 100,
             "numeric_stats": None},
            {"name": "age", "spss_label": "Âge", "stat_type_guess": "continuous",
             "n": 100, "n_missing": 1, "pct_missing": 1.0, "n_distinct": 61,
             "numeric_stats": {"min": 18, "max": 999}},
        ],
        "pii_candidates": [{"column": "tel", "kinds": ["phone"], "n_values": 100}],
    }


def test_classeur_complet_8_onglets():
    prof = _profiling_min()
    score = compute_score(CLEAN, FULL_SI)
    data = build_workbook(prof, score, SYNTH_AI)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == EXPECTED_SHEETS
    # Registre : 2 anomalies, la critique en premier (tri par gravité)
    reg = wb["Registre anomalies"]
    assert reg["A2"].value == "F001" and reg["F2"].value == "critique"
    # Variables exclues : PII 'tel' listée (déterministe)
    exclues = [r[0].value for r in wb["Variables exclues"].iter_rows(min_row=2)]
    assert "tel" in exclues
    # Résumé : le score final figure
    resume_vals = [c.value for row in wb["Résumé audit"].iter_rows() for c in row]
    assert score["score_final"] in resume_vals


def test_registre_anomalie_consolidee_par_famille():
    """Une anomalie de famille liste les variables visées et leur nombre."""
    prof = _profiling_min()
    score = compute_score(CLEAN, FULL_SI)
    ai = {"study": SYNTH_AI["study"], "dictionary": SYNTH_AI["dictionary"],
          "assessment": {**SYNTH_AI["assessment"], "findings": [
              {"id": "F001", "anomaly_class": "D", "severity": "moderee",
               "certainty": "certain", "column": None,
               "affected_columns": ["v1", "v2", "v3"], "n_affected": 3,
               "observed": None, "rule_violated": "colonnes vides",
               "title_fr": "3 variables vides", "explanation_fr": "…",
               "proposed_correction": "exclure", "requires_source_verification": False}]}}
    wb = load_workbook(io.BytesIO(build_workbook(prof, score, ai)))
    reg = wb["Registre anomalies"]
    assert reg["B2"].value == "v1, v2, v3"   # familles listées
    assert reg["C2"].value == 3              # n visés


def test_classeur_sans_audit_ia_reste_coherent():
    """Sans IA, les onglets déterministes se remplissent, les autres notent l'absence."""
    prof = _profiling_min()
    score = compute_score(CLEAN)  # sans scoring_inputs IA
    data = build_workbook(prof, score, None)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == EXPECTED_SHEETS
    # Le registre indique que l'audit IA est requis
    assert "IA" in (wb["Registre anomalies"]["A2"].value or "")
