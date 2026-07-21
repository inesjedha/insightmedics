"""Jalon M6 — Classeur Excel d'audit, au format du livrable de Hamza (10 onglets).

Assemble, à partir des sorties du pipeline (profiling, score_detail, ai_audit), le
classeur que le client reçoit : Résumé, Registre anomalies, Dictionnaire, Concordance
protocole, Journal corrections, Variables exclues, Décisions à valider, Plan analyse.
Les deux onglets de données (Base analyse, Source anonymisée) sont produits au jalon M8
(nettoyage) ; ici on écrit une note tant qu'ils ne sont pas disponibles.

100 % déterministe. Les onglets alimentés par l'IA se remplissent seulement si l'audit
IA a été exécuté ; sinon une note l'indique (le classeur reste cohérent).
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TITLE = Font(name="Arial", bold=True, size=13, color="1F3B57")
H = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HFILL = PatternFill("solid", fgColor="1F3B57")
BOLD = Font(name="Arial", bold=True, size=10)
NORM = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
SEV_FILL = {
    "critique": PatternFill("solid", fgColor="F8CBAD"),
    "critical": PatternFill("solid", fgColor="F8CBAD"),
    "majeure": PatternFill("solid", fgColor="FFE699"),
    "major": PatternFill("solid", fgColor="FFE699"),
}


def _headers(ws, headers: list[str], widths: list[int]) -> None:
    for j, (h, w) in enumerate(zip(headers, widths, strict=True), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = H
        c.fill = HFILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"


def _note(ws, text: str) -> None:
    c = ws.cell(row=2, column=1, value=text)
    c.font = Font(name="Arial", italic=True, size=10, color="888888")


# ---------------------------------------------------------------- onglets

def _sheet_resume(wb: Workbook, prof: dict, score: dict, ai: dict | None) -> None:
    ws = wb.create_sheet("Résumé audit")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    s, m = prof["structure"], prof["missing_summary"]
    ws["A1"] = "Audit biostatistique — " + prof["file"].get("file_name", "")
    ws["A1"].font = TITLE
    row = 3
    stats = [
        ("Observations (lignes)", s["n_rows"]),
        ("Variables (colonnes)", s["n_cols"]),
        ("Identifiants uniques", s["n_unique_ids"]),
        ("Colonnes entièrement vides", len(s["empty_columns"])),
        ("Taux de données manquantes", f"{m['pct_global']}%"),
        ("Doublons stricts de lignes", s["duplicates"]["exact_rows"]),
    ]
    if ai and ai.get("assessment"):
        stats.append(("Anomalies enregistrées", len(ai["assessment"].get("findings", []))))
    for label, val in stats:
        ws.cell(row=row, column=1, value=label).font = NORM
        ws.cell(row=row, column=2, value=val).font = BOLD
        row += 1
    row += 1
    # Tableau des 8 domaines
    for j, h in enumerate(["Domaine", "Score", "Maximum", "% du domaine"], start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = H
        c.fill = HFILL
    row += 1
    for d in score["domaines"]:
        ws.cell(row=row, column=1, value=f"{d['domaine']}. {d['nom']}").font = NORM
        ws.cell(row=row, column=2, value=d["obtenu"]).font = NORM
        ws.cell(row=row, column=3, value=d["max"]).font = NORM
        pct = ws.cell(row=row, column=4, value=(d["obtenu"] / d["max"] if d["max"] else 0))
        pct.number_format = "0.0%"
        pct.font = NORM
        row += 1
    ws.cell(row=row, column=1, value="Score brut").font = BOLD
    ws.cell(row=row, column=2, value=score["score_brut"]).font = BOLD
    ws.cell(row=row, column=3, value=100).font = BOLD
    row += 1
    ws.cell(row=row, column=1, value="Score officiel (après plafond)").font = BOLD
    ws.cell(row=row, column=2, value=score["score_final"]).font = BOLD
    ws.cell(row=row, column=3, value=100).font = BOLD
    row += 2
    verdict = "—"
    if ai and ai.get("assessment", {}).get("exploitability_verdict"):
        verdict = ai["assessment"]["exploitability_verdict"].get("label", "—")
    for label, val in [
        ("Verdict d'exploitabilité", verdict),
        ("Niveau de qualité", score["niveau_qualite"]),
        ("Confiance dans l'audit", score["confiance"]["niveau"]),
        ("Motif du plafond",
         "; ".join(c["defaut"] for c in score["plafonds_appliques"]) or "aucun"),
    ]:
        ws.cell(row=row, column=1, value=label).font = NORM
        ws.cell(row=row, column=2, value=val).font = BOLD
        row += 1


def _sheet_registre(wb: Workbook, ai: dict | None) -> None:
    ws = wb.create_sheet("Registre anomalies")
    ws.sheet_view.showGridLines = False
    hdr = ["id", "variable(s)", "n visés", "valeur observée", "classe (A/B/C/D)", "gravité",
           "certitude", "règle violée", "correction proposée", "vérif. dossier"]
    _headers(ws, hdr, [8, 22, 8, 16, 14, 12, 11, 28, 32, 12])
    findings = (ai or {}).get("assessment", {}).get("findings", [])
    if not findings:
        _note(ws, "Audit IA non exécuté : registre des anomalies à générer avec la clé API.")
        return
    order = {"critique": 0, "majeure": 1, "moderee": 2, "mineure": 3}
    for i, f in enumerate(sorted(findings, key=lambda x: order.get(x.get("severity"), 9)), 2):
        affected = f.get("affected_columns") or ([f["column"]] if f.get("column") else [])
        var_label = ", ".join(affected) if affected else f.get("column")
        n_aff = f.get("n_affected") or (len(affected) if affected else None)
        vals = [f.get("id"), var_label, n_aff, f.get("observed"), f.get("anomaly_class"),
                f.get("severity"), f.get("certainty"), f.get("rule_violated"),
                f.get("proposed_correction"),
                "Oui" if f.get("requires_source_verification") else "Non"]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORM
            c.alignment = WRAP
        fill = SEV_FILL.get(f.get("severity"))
        if fill:
            ws.cell(row=i, column=6).fill = fill  # colonne « gravité » (après ajout n visés)


def _sheet_dictionnaire(wb: Workbook, prof: dict, ai: dict | None) -> None:
    ws = wb.create_sheet("Dictionnaire")
    ws.sheet_view.showGridLines = False
    hdr = ["variable", "libellé", "rôle analytique", "type", "unité", "n dispo",
           "n manquant", "% manquant", "n distinct", "minimum", "maximum", "ambiguïtés"]
    _headers(ws, hdr, [18, 30, 16, 12, 10, 9, 10, 10, 10, 10, 10, 26])
    dico = {d["name"]: d for d in (ai or {}).get("dictionary", [])}
    for i, col in enumerate(prof["columns"], 2):
        d = dico.get(col["name"], {})
        ns = col.get("numeric_stats") or {}
        n = col["n"]
        vals = [
            col["name"],
            d.get("meaning") or col.get("spss_label"),
            d.get("analytic_role"),
            col.get("stat_type_guess"),
            d.get("expected_unit"),
            n - col["n_missing"],
            col["n_missing"],
            f"{col['pct_missing']}%",
            col["n_distinct"],
            ns.get("min"),
            ns.get("max"),
            d.get("ambiguities"),
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORM
            c.alignment = WRAP


def _sheet_variables_exclues(wb: Workbook, prof: dict) -> None:
    ws = wb.create_sheet("Variables exclues")
    ws.sheet_view.showGridLines = False
    _headers(ws, ["variable", "motif"], [24, 50])
    s = prof["structure"]
    rows = []
    for c in s["empty_columns"]:
        rows.append((c, "Variable entièrement vide dans la source"))
    for c in s["constant_columns"]:
        rows.append((c, "Variable constante (aucune variabilité)"))
    for p in prof.get("pii_candidates", []):
        rows.append((p["column"], "Donnée identifiante (à anonymiser avant analyse)"))
    for i, (name, motif) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=name).font = NORM
        ws.cell(row=i, column=2, value=motif).font = NORM
    if not rows:
        _note(ws, "Aucune variable à exclure détectée.")


def _sheet_journal(wb: Workbook, ai: dict | None, cleaning: dict | None = None) -> None:
    ws = wb.create_sheet("Journal corrections")
    ws.sheet_view.showGridLines = False
    # Si le nettoyage (M8) a été exécuté, on consigne les corrections RÉELLEMENT
    # appliquées à la copie ; sinon on liste les opérations PROPOSÉES par l'IA.
    if cleaning and cleaning.get("journal"):
        _headers(ws, ["id", "variable", "action appliquée (copie)", "raison",
                      "méthode", "n concerné", "vérif. dossier"],
                 [8, 26, 30, 34, 26, 11, 12])
        for i, e in enumerate(cleaning["journal"], 2):
            vals = [e.get("id_correction"), e.get("variable"),
                    e.get("action_appliquee_dans_copie"), e.get("raison"),
                    e.get("methode"), e.get("n_concerne"),
                    e.get("verification_dossier_necessaire")]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=i, column=j, value=v)
                c.font = NORM
                c.alignment = WRAP
        return
    _headers(ws, ["id", "opération", "variable", "justification", "auto-applicable"],
             [8, 22, 18, 40, 14])
    plan = (ai or {}).get("assessment", {}).get("cleaning_plan", [])
    if not plan:
        _note(ws, "Plan de nettoyage proposé disponible après l'audit IA "
                  "(appliqué à la copie dérivée au jalon M8).")
        return
    for i, op in enumerate(plan, 2):
        vals = [op.get("op_id"), op.get("operation"), op.get("column"),
                op.get("rationale_fr"), "Oui" if op.get("auto_safe") else "Validation requise"]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORM
            c.alignment = WRAP


def _sheet_decisions(wb: Workbook, ai: dict | None) -> None:
    ws = wb.create_sheet("Décisions à valider")
    ws.sheet_view.showGridLines = False
    _headers(ws, ["question", "variable", "options", "conséquence", "recommandation"],
             [40, 16, 24, 32, 32])
    decs = (ai or {}).get("assessment", {}).get("client_decisions", [])
    if not decs:
        _note(ws, "Décisions à valider produites avec l'audit IA.")
        return
    for i, d in enumerate(decs, 2):
        vals = [d.get("question_fr"), d.get("column"), " / ".join(d.get("options", [])),
                d.get("consequence_fr"), d.get("recommendation_fr")]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORM
            c.alignment = WRAP


def _sheet_concordance(wb: Workbook, ai: dict | None) -> None:
    ws = wb.create_sheet("Concordance protocole")
    ws.sheet_view.showGridLines = False
    _headers(ws, ["objectif", "critère de jugement", "variables candidates", "faisabilité"],
             [40, 30, 30, 20])
    study = (ai or {}).get("study")
    if not study:
        _note(ws, "Concordance protocole disponible si un protocole a été fourni.")
        return
    ep = study.get("primary_endpoint") or {}
    ws.cell(row=2, column=1, value=study.get("primary_objective")).font = NORM
    ws.cell(row=2, column=2, value=ep.get("description")).font = NORM
    ws.cell(row=2, column=3, value=", ".join(ep.get("candidate_columns", []))).font = NORM
    for r, obj in enumerate(study.get("secondary_objectives", []), 3):
        ws.cell(row=r, column=1, value=obj).font = NORM
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP


def _sheet_plan(wb: Workbook, ai: dict | None) -> None:
    ws = wb.create_sheet("Plan analyse")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    rs = (ai or {}).get("assessment", {}).get("report_sections_fr", {})
    plan = rs.get("plan_analyse_conditionnel")
    if not plan:
        ws["A1"] = "Plan d'analyse conditionnel disponible avec l'audit IA."
        ws["A1"].font = Font(name="Arial", italic=True, color="888888")
        return
    ws["A1"] = "Plan d'analyse statistique conditionnel"
    ws["A1"].font = TITLE
    c = ws.cell(row=3, column=1, value=plan)
    c.font = NORM
    c.alignment = WRAP


def _sheet_data(wb: Workbook, title: str, df, note: str, max_rows: int = 500) -> None:
    """Écrit un DataFrame (base nettoyée / source anonymisée) en onglet, entêtes figées."""
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    cols = list(df.columns)
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=j, value=str(name))
        c.font = H
        c.fill = HFILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, len(str(name)) + 2), 24)
    ws.freeze_panes = "A2"
    shown = df.head(max_rows)
    for i, (_, r) in enumerate(shown.iterrows(), 2):
        for j, name in enumerate(cols, 1):
            v = r[name]
            try:
                import pandas as _pd
                if _pd.isna(v):
                    v = None
            except Exception:  # noqa: BLE001
                pass
            ws.cell(row=i, column=j, value=v).font = NORM
    if len(df) > max_rows:
        r = len(shown) + 3
        ws.cell(row=r, column=1,
                value=f"… {len(df) - max_rows} ligne(s) supplémentaire(s). "
                      f"Base complète dans les fichiers .sav / .csv joints.").font = \
            Font(name="Arial", italic=True, size=10, color="888888")
    elif note:
        pass


# ---------------------------------------------------------------- entrée

def build_workbook(profiling: dict, score_detail: dict,
                   ai_audit: dict | None, cleaning: dict | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # retire la feuille vide par défaut
    _sheet_resume(wb, profiling, score_detail, ai_audit)
    _sheet_registre(wb, ai_audit)
    _sheet_dictionnaire(wb, profiling, ai_audit)
    _sheet_concordance(wb, ai_audit)
    _sheet_journal(wb, ai_audit, cleaning)
    _sheet_variables_exclues(wb, profiling)
    _sheet_decisions(wb, ai_audit)
    _sheet_plan(wb, ai_audit)
    # Onglets de données produits au jalon M8 (nettoyage). Ajoutés seulement si dispo.
    if cleaning:
        _sheet_data(wb, "Base analyse", cleaning["base_analyse"],
                    "Copie nettoyée et anonymisée ; original inchangé.")
        _sheet_data(wb, "Source anonymisée", cleaning["source_anonymisee"],
                    "Toutes les colonnes, identifiants directs retirés.")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
